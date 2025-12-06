"""Excel服务 - 生成财务凭证Excel表格"""
import io
from datetime import datetime
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# Excel表头定义
EXCEL_HEADERS = [
    "编制日期",
    "凭证类型",
    "凭证序号",
    "凭证号",
    "制单人",
    "附件张数",
    "会计年度",
    "科目编码",
    "科目名称",
    "凭证摘要",
    "借贷方向",
    "金额",
    "币种",
    "汇率",
    "原币金额",
    "数量",
    "单价",
    "结算方式名称",
    "结算日期",
    "结算票号",
    "业务日期",
    "员工编号",
    "员工姓名",
    "往来单位编号",
    "往来单位名称",
    "货品编号",
    "货品名称",
    "部门名称",
    "项目名称",
]

# 字段映射（从LLM返回的数据字段到Excel列）
FIELD_MAPPING = {
    "voucher_date": "编制日期",
    "voucher_type": "凭证类型",
    "voucher_seq": "凭证序号",
    "voucher_no": "凭证号",
    "preparer": "制单人",
    "attachment_count": "附件张数",
    "fiscal_year": "会计年度",
    "subject_code": "科目编码",
    "subject_name": "科目名称",
    "summary": "凭证摘要",
    "direction": "借贷方向",
    "amount": "金额",
    "currency": "币种",
    "exchange_rate": "汇率",
    "original_amount": "原币金额",
    "quantity": "数量",
    "unit_price": "单价",
    "settlement_method": "结算方式名称",
    "settlement_date": "结算日期",
    "settlement_no": "结算票号",
    "business_date": "业务日期",
    "employee_no": "员工编号",
    "employee_name": "员工姓名",
    "partner_no": "往来单位编号",
    "partner_name": "往来单位名称",
    "product_no": "货品编号",
    "product_name": "货品名称",
    "department": "部门名称",
    "project": "项目名称",
}


class ExcelService:
    """Excel服务"""
    
    def __init__(self):
        self.headers = EXCEL_HEADERS
        self.field_mapping = FIELD_MAPPING
    
    def create_workbook(self) -> Workbook:
        """创建新的工作簿"""
        wb = Workbook()
        ws = wb.active
        ws.title = "财务凭证"
        
        # 设置表头样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # 写入表头
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 设置列宽
        column_widths = {
            "编制日期": 12,
            "凭证类型": 10,
            "凭证序号": 8,
            "凭证号": 10,
            "制单人": 10,
            "附件张数": 8,
            "会计年度": 10,
            "科目编码": 10,
            "科目名称": 15,
            "凭证摘要": 20,
            "借贷方向": 8,
            "金额": 12,
            "币种": 8,
            "汇率": 8,
            "原币金额": 12,
            "数量": 10,
            "单价": 10,
            "结算方式名称": 12,
            "结算日期": 12,
            "结算票号": 12,
            "业务日期": 12,
            "员工编号": 10,
            "员工姓名": 10,
            "往来单位编号": 12,
            "往来单位名称": 15,
            "货品编号": 10,
            "货品名称": 15,
            "部门名称": 12,
            "项目名称": 15,
        }
        
        for col, header in enumerate(self.headers, 1):
            ws.column_dimensions[get_column_letter(col)].width = column_widths.get(header, 12)
        
        # 冻结首行
        ws.freeze_panes = "A2"
        
        return wb
    
    def add_voucher_data(self, ws, voucher_data: dict, voucher_seq: int = 1):
        """添加凭证数据到工作表"""
        # 获取当前最后一行
        last_row = ws.max_row
        
        # 通用凭证信息
        common_data = {
            "编制日期": voucher_data.get("voucher_date", ""),
            "凭证类型": voucher_data.get("voucher_type", "记"),
            "凭证序号": voucher_seq,
            "凭证号": voucher_data.get("voucher_no", ""),
            "制单人": voucher_data.get("preparer", ""),
            "附件张数": voucher_data.get("attachment_count", 0),
            "会计年度": voucher_data.get("fiscal_year", ""),
        }
        
        # 设置单元格样式
        data_alignment = Alignment(horizontal="left", vertical="center")
        number_alignment = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # 处理每条分录
        entries = voucher_data.get("entries", [])
        if not entries:
            entries = [{}]  # 至少添加一行空数据
        
        for entry in entries:
            last_row += 1
            row_data = common_data.copy()
            
            # 添加分录数据
            row_data.update({
                "科目编码": entry.get("subject_code", ""),
                "科目名称": entry.get("subject_name", ""),
                "凭证摘要": entry.get("summary", ""),
                "借贷方向": entry.get("direction", ""),
                "金额": entry.get("amount", ""),
                "币种": entry.get("currency", "人民币"),
                "汇率": entry.get("exchange_rate", 1),
                "原币金额": entry.get("original_amount", ""),
                "数量": entry.get("quantity", ""),
                "单价": entry.get("unit_price", ""),
                "结算方式名称": entry.get("settlement_method", ""),
                "结算日期": entry.get("settlement_date", ""),
                "结算票号": entry.get("settlement_no", ""),
                "业务日期": entry.get("business_date", ""),
                "员工编号": entry.get("employee_no", ""),
                "员工姓名": entry.get("employee_name", ""),
                "往来单位编号": entry.get("partner_no", ""),
                "往来单位名称": entry.get("partner_name", ""),
                "货品编号": entry.get("product_no", ""),
                "货品名称": entry.get("product_name", ""),
                "部门名称": entry.get("department", ""),
                "项目名称": entry.get("project", ""),
            })
            
            # 写入数据
            for col, header in enumerate(self.headers, 1):
                value = row_data.get(header, "")
                cell = ws.cell(row=last_row, column=col, value=value)
                cell.border = thin_border
                
                # 数字列右对齐
                if header in ["金额", "原币金额", "数量", "单价", "汇率", "附件张数", "凭证序号"]:
                    cell.alignment = number_alignment
                else:
                    cell.alignment = data_alignment
    
    def generate_excel(self, vouchers: list[dict]) -> bytes:
        """
        生成Excel文件
        
        Args:
            vouchers: 凭证数据列表，每个元素是一个凭证的结构化数据
        
        Returns:
            Excel文件的字节数据
        """
        wb = self.create_workbook()
        ws = wb.active
        
        # 添加所有凭证数据
        for seq, voucher in enumerate(vouchers, 1):
            if "error" not in voucher:
                self.add_voucher_data(ws, voucher, seq)
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def generate_template(self) -> bytes:
        """生成空白模板"""
        wb = self.create_workbook()
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()

